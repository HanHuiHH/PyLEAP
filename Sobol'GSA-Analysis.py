# -*- coding:utf-8 -*-
# @Time : 2023-2-23 22:40
# @Author: Hui Han
# @File : Sobol-LEAP
import os
import shutil
import time

import openpyxl as op
import pandas
from SALib.analyze import sobol
import numpy as np
import matplotlib.pyplot as plt

config = {
    "font.family": 'serif',
    "font.size": 12,  # 相当于小四字体
    "font.serif": ['SongNTR'],  # 宋体
    'axes.unicode_minus': False,  # 处理负号，即-号
    "figure.figsize": (16, 6)
}
plt.rcParams.update(config)
start_time = time.time()

# TODO:定义需要进行敏感性分析的参数的信息
problem = {
    'num_vars': 8,  # 需要测试敏感性的参数个数，可能和模型输入个数不同
    'names': ["PerCapGDP", "IndProp", "PopGrow", "ElecSave", "EVProp", "PVCapacity", "ImpElec", "Hydrogen"],
    # 需要测试敏感性的参数名称
    'dists': ['norm', 'norm', 'norm', 'norm', 'norm', 'norm', 'norm', 'norm'],
    'bounds': [[0.5, 0.2], [0.5, 0.2], [0.5, 0.2], [0.5, 0.2], [0.5, 0.2], [0.5, 0.2], [0.5, 0.2], [0.5, 0.2]]
    # 定义每个参数的上下界
}


# TODO:导入LEAP蒙特卡洛模拟的结果
files = os.listdir("./Results/Simulation")
# print(files)
file_name = "./Results/Simulation/" + files[-1]
# file_name = "E:/1安徽碳中和/3模型计算/LEAP情景组合/模型记录/"
print('Reading file:', file_name)
df_dict = pandas.read_excel(file_name, sheet_name=None, header=0)
data_list = []
data_name = []
for df in df_dict:
    data = pandas.read_excel(file_name, sheet_name=df, header=0)
    data_list.append(data)
    data_name.append(df)
# print(data_list)

# TODO:将要分析的导入结果设成y
analysis_col = 2
y = np.array(data_list[analysis_col])  # 通过模型运行获得生成的结果y
"""y在敏感性分析前就生成了，说明可以依据采样的参数输入模型跑出结果后，再进行敏感性分析"""

# analyse
sobol_indices = [sobol.analyze(problem, Y, print_to_console=True) for Y in y.T]  # sobol分析每个x取值上的a、b敏感性大小，所以长度等于x的个数


def save_sobol_result(sobol_indices):
    # TODO:保存sobol分析参数结果至excel
    S1 = []
    S1_conf = []
    ST = []
    ST_conf = []
    for indice in sobol_indices:
        S1.append(np.array(indice['S1']))
        S1_conf.append(np.array(indice['S1_conf']))
        ST.append(np.array(indice['ST']))
        ST_conf.append(np.array(indice['ST_conf']))
    S2_2030 = sobol_indices[11]['S2']
    S2_2030_conf = sobol_indices[11]['S2_conf']
    S2_2060 = sobol_indices[-1]['S2']
    S2_2060_conf = sobol_indices[-1]['S2_conf']

    # TODO：修改sobol分析参数结果保存路径
    ResultPath = "./Results/SobolResults/SobolResults" + \
                 time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime()) + ".xlsx"
    shutil.copyfile("ImportExcel/3 Sobol Results (blank).xlsx", ResultPath)
    wb = op.load_workbook(filename=ResultPath)
    ws = wb["ST"]
    for i in range(42):
        for j in range(8):
            ws.cell(row=i + 2, column=j + 2, value=ST[i][j])
    ws = wb["ST_conf"]
    for i in range(42):
        for j in range(8):
            ws.cell(row=i + 2, column=j + 2, value=ST_conf[i][j])
    ws = wb["S1"]
    for i in range(42):
        for j in range(8):
            ws.cell(row=i + 2, column=j + 2, value=S1[i][j])
    ws = wb["S1_conf"]
    for i in range(42):
        for j in range(8):
            ws.cell(row=i + 2, column=j + 2, value=S1_conf[i][j])
    ws = wb["S2_2030"]
    for i in range(8):
        for j in range(8):
            ws.cell(row=i + 2, column=j + 2, value=S2_2030[i][j])
    ws = wb["S2_2030_conf"]
    for i in range(8):
        for j in range(8):
            ws.cell(row=i + 2, column=j + 2, value=S2_2030_conf[i][j])
    ws = wb["S2_2060"]
    for i in range(8):
        for j in range(8):
            ws.cell(row=i + 2, column=j + 2, value=S2_2060[i][j])
    ws = wb["S2_2060_conf"]
    for i in range(8):
        for j in range(8):
            ws.cell(row=i + 2, column=j + 2, value=S2_2060_conf[i][j])
    wb.save(filename=ResultPath)


save_sobol_result(sobol_indices)

# '''
#     作热力图
# '''
# heatmap(sobol_indices[11]['S2'])
# plt.show()
# heatmap(sobol_indices[-1]['S2'])
# plt.show()


# TODO:开始对结果进行作图
# '''
#     一阶敏感性
# '''
# S1s = np.array([s['S1'] for s in sobol_indices])  # 每个x取值采样对应的a和b的一阶敏感性系数
#
# fig = plt.figure(constrained_layout=True)
#
# gs = fig.add_gridspec(2, 5)
#
# ax_list = [fig.add_subplot(gs[:, 0])]  # The first plot draw in the left is Y
# for i in range((problem['num_vars'] + 1) // 2):  # Other plots draw S1 of params, number in range() must be integer
#     ax_list.append(fig.add_subplot(gs[0, i + 1]))  # Add a plot in upper part
#     ax_list.append(fig.add_subplot(gs[1, i + 1]))  # Add a plot in lower part
# if problem['num_vars'] % 2 == 1:  # If the number of vars is not even, del the last plot
#     del ax_list[-1]
#
# years = list(range(2020, 2061))
# for i, ax in enumerate(ax_list[1:]):  # Draw all plots but the first
#     ax.plot(years, S1s[1:, i],
#             label=r'S1$_\mathregular{{{}}}$'.format(problem["names"][i]),
#             color='black')
#     ax.set_ylabel("First-order Sobol index")
#
#     ax.set_ylim(0, 0.5)
#
#     ax.yaxis.set_label_position("left")
#     ax.yaxis.tick_left()
#
#     ax.legend(loc='upper left', fontsize=16)
#
# ax_list[0].plot(years, np.mean(y[:, 1:], axis=0), label="Mean", color='black')
#
# # in percent
# prediction_interval = 95
# # 简单填充
# ax_list[0].fill_between(years,
#                         np.percentile(y[:, 1:], 50 - prediction_interval / 2., axis=0),
#                         # np.percentile函数获取2.5%和97.5%的大小上下界
#                         np.percentile(y[:, 1:], 50 + prediction_interval / 2., axis=0),
#                         alpha=0.5, color='black',
#                         label=f"{prediction_interval} % prediction interval")
#
# ax_list[0].set_xlabel("x")
# ax_list[0].set_ylabel("y")
# ax_list[0].legend(title=data_name[analysis_col], loc='upper center')._legend_box.align = "left"
#
# plt.savefig('E:/1安徽碳中和/3模型计算/LEAP情景组合/sobol分析作图/S1'
#             + time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime()) + '.jpg', dpi=600)
# plt.show()

'''
    总体敏感性
'''
STs = np.array([s['ST'] for s in sobol_indices])  # 每个x取值采样对应的a和b的一阶敏感性系数
S1s = np.array([s['S1'] for s in sobol_indices])  # 每个x取值采样对应的a和b的一阶敏感性系数

# fig = plt.figure(constrained_layout=True)
fig = plt.figure()

gs = fig.add_gridspec(2, 4)

ax_list = []  # Create a list of fig list
for i in range((problem['num_vars'] + 1) // 2):  # Other plots draw S1 of params, number in range() must be integer
    ax_list.append(fig.add_subplot(gs[0, i]))  # Add a plot in upper part
    ax_list.append(fig.add_subplot(gs[1, i]))  # Add a plot in lower part
if problem['num_vars'] % 2 == 1:  # If the number of vars is not even, del the last plot
    del ax_list[-1]

# Draw the curves of total sensitivity in blue colour
years = list(range(2020, 2061))
for i, ax in enumerate(ax_list):  # Draw all plots but the first
    if i == 0:
        ax.plot(years, STs[1:, i],
                label='Total sensitivity',
                color='blue')
    else:
        ax.plot(years, STs[1:, i],
                color='blue')
    ax.set_ylabel("Sobol' index", fontsize=16)

    # ax.set_ylim(0, 0.4)
    ax.set_ylim(0, )

    ax.yaxis.set_label_position("left")
    ax.yaxis.tick_left()

    # ax.legend(loc='upper left', fontsize=16)
    ax.set_title("({}) ".format(chr(i + 97)) + problem["names"][i])  # Set title of each graph as "(n) Name"

# Draw the curves of total sensitivity in gray colour
for i, ax in enumerate(ax_list):  # Draw all plots but the first
    if i == 0:
        ax.plot(years, S1s[1:, i],
                label='First-order sensitivity',
                color='gray')
    else:
        ax.plot(years, S1s[1:, i],
                color='gray')

    ax.set_ylim(0, )

    ax.yaxis.set_label_position("left")
    ax.yaxis.tick_left()

for ax in ax_list[2:]:
    ax.set_ylabel('')  # Just keep the left y-axis label

fig.legend(loc='upper center', ncol=4, bbox_to_anchor=(0.5, 1.01), fontsize=16)  # Add legend
fig.subplots_adjust(hspace=0.3, wspace=0.25)  # Adjust location
plt.savefig('./Results/Figures/ST' + time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime()) + '.jpg', dpi=600)
plt.show()

#
# '''
# Function to convert S2 output for graph generation. Taken from
# https://github.com/Project-Platypus/Rhodium/blob/master/rhodium/sa.py
# '''
# # Todo: Change number in [] to choose a year of result to draw pictures
# #       11 is 2030 (carbon peak), -1 is 2060 (carbon neutral)
# Si = sobol_indices[11]
#
#
# def S2_to_dict(matrix, problem):
#     result = {}
#     names = list(problem["names"])
#
#     for i in range(problem["num_vars"]):
#         for j in range(i + 1, problem["num_vars"]):
#             if names[i] not in result:
#                 result[names[i]] = {}
#             if names[j] not in result:
#                 result[names[j]] = {}
#
#             result[names[i]][names[j]] = result[names[j]][names[i]] = float(matrix[i][j])
#
#     return result
#
#
# result = {'S1': {k: float(v) for k, v in zip(problem["names"], Si["S1"])},
#           'S1_conf': {k: float(v) for k, v in zip(problem["names"], Si["S1_conf"])},
#           'S2': S2_to_dict(Si['S2'], problem),
#           'S2_conf': S2_to_dict(Si['S2_conf'], problem),
#           'ST': {k: float(v) for k, v in zip(problem["names"], Si["ST"])},
#           'ST_conf': {k: float(v) for k, v in zip(problem["names"], Si["ST_conf"])}}  # create dictionary to store new
#
# drawgraphs(result)
