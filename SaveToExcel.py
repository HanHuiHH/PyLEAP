"""
    本代码封装保存模拟结果功能
    无法单独使用
"""

import time
import openpyxl as op


def save_all_records(AllRecords, ResultPath):
    """
    输出所有模拟结果到预先编辑好的excel
    :param AllRecords: 所有结果参数的记录
    :param ResultPath: 保存的位置
    :return: Nothing
    """
    wb = op.load_workbook(filename=ResultPath)
    ws = wb["能源消费记录"]
    for i in range(len(AllRecords["EnergyDataRecord"])):
        for j in range(len(AllRecords["EnergyDataRecord"][i])):
            ws.cell(row=i + 2, column=j + 1, value=AllRecords["EnergyDataRecord"][i][j])
    ws = wb["碳排放记录"]
    for i in range(len(AllRecords["EmissionDataRecord"])):
        for j in range(len(AllRecords["EmissionDataRecord"][i])):
            ws.cell(row=i + 2, column=j + 1, value=AllRecords["EmissionDataRecord"][i][j])
    ws = wb["净碳排放记录"]
    for i in range(len(AllRecords["NetEmissionRecord"])):
        for j in range(len(AllRecords["NetEmissionRecord"][i])):
            ws.cell(row=i + 2, column=j + 1, value=AllRecords["NetEmissionRecord"][i][j])
    ws = wb["电力消费记录"]
    for i in range(len(AllRecords["ElecConsDataRecord"])):
        for j in range(len(AllRecords["ElecConsDataRecord"][i])):
            ws.cell(row=i + 2, column=j + 1, value=AllRecords["ElecConsDataRecord"][i][j])
    ws = wb["清洁能源占比记录"]
    for i in range(len(AllRecords["CleanPropRecord"])):
        for j in range(len(AllRecords["CleanPropRecord"][i])):
            ws.cell(row=i + 2, column=j + 1,
                    value=AllRecords["CleanPropRecord"][i][j]).number_format = '0.00%'  # 设置为百分比格式
    ws = wb["能耗强度下降率记录"]
    for i in range(len(AllRecords["EnergyIntensityRecord"])):
        for j in range(len(AllRecords["EnergyIntensityRecord"][i])):
            ws.cell(row=i + 2, column=j + 1,
                    value=AllRecords["EnergyIntensityRecord"][i][j]).number_format = '0.00%'  # 设置为百分比格式
    ws = wb["碳排放强度下降率记录"]
    for i in range(len(AllRecords["CarbonIntensityRecord"])):
        for j in range(len(AllRecords["CarbonIntensityRecord"][i])):
            ws.cell(row=i + 2, column=j + 1,
                    value=AllRecords["CarbonIntensityRecord"][i][j]).number_format = '0.00%'  # 设置为百分比格式
    wb.save(filename=ResultPath)

