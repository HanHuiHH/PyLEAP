# -*- coding:utf-8 -*-
# @Time : 2023-2-23 22:40
# @Author: Hui Han
# @File : Sobol-LEAP

import time
from SALib.sample import sobol as sobol_sample
from SALib.analyze import sobol as sobol_analyze
import numpy as np
import matplotlib.pyplot as plt

from Simulation import test, InitialAllRecords
import shutil
import win32com.client as client
import openpyxl as op
from ImportFromExcel import GetValueRange
from tqdm import *


def save_all_records(AllRecords, ResultPath):
    """⑦输出所有模拟结果到excel"""
    wb = op.load_workbook(filename=ResultPath)
    ws = wb["能源消费记录"]
    for i in range(len(AllRecords["EnergyDataRecord"])):
        for j in range(len(AllRecords["EnergyDataRecord"][i])):
            ws.cell(row=i + 2, column=j + 1, value=AllRecords["EnergyDataRecord"][i][j])
    ws = wb["碳排放记录"]
    for i in range(len(AllRecords["EmissionDataRecord"])):
        for j in range(len(AllRecords["EmissionDataRecord"][i])):
            ws.cell(row=i + 2, column=j + 1, value=AllRecords["EmissionDataRecord"][i][j])
    ws = wb["净碳排放记录"]
    for i in range(len(AllRecords["NetEmissionRecord"])):
        for j in range(len(AllRecords["NetEmissionRecord"][i])):
            ws.cell(row=i + 2, column=j + 1, value=AllRecords["NetEmissionRecord"][i][j])
    ws = wb["电力消费记录"]
    for i in range(len(AllRecords["ElecConsDataRecord"])):
        for j in range(len(AllRecords["ElecConsDataRecord"][i])):
            ws.cell(row=i + 2, column=j + 1, value=AllRecords["ElecConsDataRecord"][i][j])
    ws = wb["清洁能源占比记录"]
    for i in range(len(AllRecords["CleanPropRecord"])):
        for j in range(len(AllRecords["CleanPropRecord"][i])):
            if j > len(MySimuCount):  # 非PI值需要设置为百分比格式
                ws.cell(row=i + 2, column=j + 1, value=AllRecords["CleanPropRecord"][i][j]).number_format = '0.00%'
            else:
                ws.cell(row=i + 2, column=j + 1, value=AllRecords["CleanPropRecord"][i][j])
    ws = wb["能耗强度下降率记录"]
    for i in range(len(AllRecords["EnergyIntensityRecord"])):
        for j in range(len(AllRecords["EnergyIntensityRecord"][i])):
            if j > len(MySimuCount):  # 非PI值需要设置为百分比格式
                ws.cell(row=i + 2, column=j + 1,
                        value=AllRecords["EnergyIntensityRecord"][i][j]).number_format = '0.00%'
            else:
                ws.cell(row=i + 2, column=j + 1, value=AllRecords["EnergyIntensityRecord"][i][j])
    ws = wb["碳排放强度下降率记录"]
    for i in range(len(AllRecords["CarbonIntensityRecord"])):
        for j in range(len(AllRecords["CarbonIntensityRecord"][i])):
            if j > len(MySimuCount):  # 非PI值需要设置为百分比格式
                ws.cell(row=i + 2, column=j + 1,
                        value=AllRecords["CarbonIntensityRecord"][i][j]).number_format = '0.00%'
            else:
                ws.cell(row=i + 2, column=j + 1, value=AllRecords["CarbonIntensityRecord"][i][j])
    wb.save(filename=ResultPath)


start_time = time.time()

leap = client.DispatchEx('leap.LEAPApplication')  # 启动独立的进程
leap.Visible = 0  # 0表示在后台以进程方式运行，不显示软件界面，1表示显示软件界面并可能需要操作
MySimuCount = [3, 3, 3, 3, 3, 3, 3, 3]  # TODO:设置每个parameter的变化情况,不变就设置成1，变化的设置成变化维度

ResultPath = "E:/1安徽碳中和/3模型计算/LEAP情景组合/sobol分析记录/sobol分析记录" + \
             time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime()) + ".xlsx"  # TODO：修改结果保存路径
shutil.copyfile(r"E:\1安徽碳中和\3模型计算\LEAP情景组合\sobol分析空白记录.xlsx", ResultPath)  # TODO：空白参数记录文件地址
OriginAreaName = "20230409安徽省碳排放总模型"

# TODO:定义需要进行敏感性分析的参数的信息
problem = {
    'num_vars': 8,  # 需要测试敏感性的参数个数，可能和模型输入个数不同
    'names': ["GDP", "IndPorp", "PopGrow", "ElecSave", "EVProp", "PVGrow", "ImpElecGrow", "HydrogenRep"],  # 需要测试敏感性的参数名称
    'dists': ['norm', 'norm', 'norm', 'norm', 'norm', 'norm', 'norm', 'norm'],
    'bounds': [[0.5, 0.2], [0.5, 0.2], [0.5, 0.2], [0.5, 0.2], [0.5, 0.2], [0.5, 0.2], [0.5, 0.2], [0.5, 0.2]]  # 定义每个参数的均值和方差
}

# TODO:开始进行敏感性分析
# sample
N_value = 2 ** 8  # 敏感性分析的N
param_values = sobol_sample.sample(problem, N_value)  # 后一个是敏感性分析N，sobol分析次数为N*(D+1)，D为分析维度，N推荐为500~1000

# evaluate
NewAreaName = "Simulation0"  # 定义模拟Area的名称为Simulation
leap.Areas.Add(NewAreaName, OriginAreaName)  # 创建一个新的LEAP Area，如果报错，打开任务管理器把所有的LEAP进程都关掉重新试一试
leap.Areas(NewAreaName).Open()
leap.Scenarios("Comprehensive Scenario").Active = True  # 定位到所需要的情景
print("目前打开的Area：", leap.ActiveArea.Name)
print("目前打开的Scenario：", leap.ActiveScenario.Name)  # 如果scenario变成current account，在LEAP里面选择综合情景再直接模拟
AllRecords = InitialAllRecords()

"""（2）提前进行 氢能替代 参数设置，否则会重复累加修改leap中参数"""
HyIndexToChange = ["钢铁", "有色", "化工", "建材", "其他"]
HyOriginExpression5th = []
for Indi in range(len(HyIndexToChange)):  # 先保存原来的表达式，以免堆积修改
    Index = r"Demand\Industry\ " + str(HyIndexToChange[Indi]) + "\Hydrogen"
    HyOriginExpression5th.append(leap.Branch(Index).Variable("Final Energy Intensity").Expression)
"""（3）提前进行 用电节能 参数设置，否则会重复累加修改leap中参数"""
IndIndexToChange = ["钢铁", "有色", "化工", "建材", "其他"]
SerIndexToChange = ["Wholesale", "Public Service", "IT Service", "Real Estate", "Financial Service"]
IndOriginExpression6th = []
SerOrrginExpression6th = []
for Indi in range(len(IndIndexToChange)):  # 先保存原来的表达式，以免堆积修改
    IndIndex = r"Demand\Industry\ " + str(IndIndexToChange[Indi]) + "\Electricity"
    IndOriginExpression6th.append(
        leap.Branch(IndIndex).Variable("Final Energy Intensity").Expression
    )
for Seri in range(len(SerIndexToChange)):  # 先保存原来的表达式，以免堆积修改
    SerIndex = r"Demand\OtherService\ " + str(SerIndexToChange[Seri]) + "\Electricity"
    SerOrrginExpression6th.append(
        leap.Branch(SerIndex).Variable("Final Energy Intensity").Expression
    )
saved_origin_parameter = [IndIndexToChange, IndOriginExpression6th, SerIndexToChange, SerOrrginExpression6th,
                          HyIndexToChange, HyOriginExpression5th]

for params in tqdm(param_values):
    AllRecords = test(
        leap=leap,
        AllRanges=GetValueRange(FilePath=r"E:\1安徽碳中和\3模型计算\LEAP情景组合\重点模拟参数.xlsx",
                                SheetName="Sheet2", SimuCount=MySimuCount),
        AllRecords=AllRecords,
        params=params,
        saved_origin_parameter=saved_origin_parameter
    )
    if len(AllRecords["EnergyDataRecord"]) % (8 * 2 * (8 + 1)) == 0:  # 保存次数应该是N*（2D+2），否则sobol分析会报错
        save_all_records(AllRecords, ResultPath)
        print("===================保存文件，已完成{}次模拟===================\n".format(len(AllRecords["EnergyDataRecord"])))

save_all_records(AllRecords, ResultPath)  # 所有模拟完成后最后保存一次文件

"""⑥模拟结束，关闭并删除模拟使用的Area，否则重新模拟会报错"""
# leap.Areas(NewAreaName).Save()
leap.Areas(OriginAreaName).Open()
leap.Areas.Delete(NewAreaName)
leap.Visible = 1  # 显示LEAP界面，便于关闭