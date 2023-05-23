"""
    本代码封装从excel中读取变量参数的功能
    无法单独使用
"""

import time
import openpyxl as op


def GetValueRange(FilePath, SheetName):
    """
    从excel中读取变量参数
    输入指标文件路径、表格名称、各参数模拟的次数，输出指标名称和对应范围列表的字典
    :param FilePath:指标excel文件路径
    :param SheetName:表格名称
    :return:指标名称和对应范围列表的字典
    """
    wb = op.load_workbook(filename=FilePath, data_only=True)
    ws = wb[SheetName]

    row_accumulated = 2
    GDPRanges = []  # Range为该指标多个Range的集合（3×9），该指标为人均GDP增长率
    for i in range(3):  # range(3)指这个指标分了3个层次，分别为高值、中值和低值，如果细分更多层次需要修改
        row = row_accumulated + i
        GDPRanges.append([])  # 每个层次创建独立的列表进行储存
        for j in range(9):
            GDPRanges[i].append(ws[chr(ord("E") + j) + str(row)].value)  # 从2020年开始填入列表值
    row_accumulated += 3

    IndPorpRanges = []  # Range为该指标多个Range的集合（3×9），该指标为工业占比
    for i in range(3):  # range(3)指这个指标分了3个层次，如果细分更多层次需要修改
        row = row_accumulated + i
        IndPorpRanges.append([])  # 每个层次创建独立的列表进行储存
        for j in range(9):
            IndPorpRanges[i].append(ws[chr(ord("E") + j) + str(row)].value)  # 从2020年开始填入列表值
    row_accumulated += 3

    PopGrowRanges = []  # Range为该指标多个Range的集合（3×9），该指标为人口增长率
    for i in range(3):  # range(3)指这个指标分了3个层次，如果细分更多层次需要修改
        row = row_accumulated + i
        PopGrowRanges.append([])  # 每个层次创建独立的列表进行储存
        for j in range(9):
            PopGrowRanges[i].append(ws[chr(ord("E") + j) + str(row)].value)  # 从2020年开始填入列表值
    row_accumulated += 3

    ElecSaveRanges = []  # Range为该指标多个Range的集合（3×9），该指标为用电节能速度
    for i in range(3):  # range(3)指这个指标分了3个层次，如果细分更多层次需要修改
        row = row_accumulated + i
        ElecSaveRanges.append([])  # 每个层次创建独立的列表进行储存
        for j in range(9):
            ElecSaveRanges[i].append(ws[chr(ord("E") + j) + str(row)].value)  # 从2020年开始填入列表值
    row_accumulated += 3

    EVPropRanges = []  # Range为该指标多个Range的集合（3×9），该指标为电动车占比
    for i in range(3):  # range(3)指这个指标分了3个层次，如果细分更多层次需要修改
        row = row_accumulated + i
        EVPropRanges.append([])  # 每个层次创建独立的列表进行储存
        for j in range(9):
            EVPropRanges[i].append(ws[chr(ord("E") + j) + str(row)].value)  # 从2020年开始填入列表值
    row_accumulated += 3

    PVGrowRanges = []  # Range为该指标多个Range的集合（3×9），该指标为光伏发电增长率
    for i in range(3):  # range(3)指这个指标分了3个层次，如果细分更多层次需要修改
        row = row_accumulated + i
        PVGrowRanges.append([])  # 每个层次创建独立的列表进行储存
        for j in range(9):
            PVGrowRanges[i].append(ws[chr(ord("E") + j) + str(row)].value)  # 从2020年开始填入列表值
    row_accumulated += 3

    ImpElecGrowRanges = []  # Range为该指标多个Range的集合（3×9），该指标为调入电力量
    for i in range(3):  # range(3)指这个指标分了3个层次，如果细分更多层次需要修改
        row = row_accumulated + i
        ImpElecGrowRanges.append([])  # 每个层次创建独立的列表进行储存
        for j in range(9):
            ImpElecGrowRanges[i].append(ws[chr(ord("E") + j) + str(row)].value)  # 从2020年开始填入列表值
    row_accumulated += 3

    HydrogenRepRanges = []  # Range为该指标多个Range的集合（3×9），该指标为氢能替代速率
    for i in range(3):  # range(3)指这个指标分了3个层次，如果细分更多层次需要修改
        row = row_accumulated + i
        HydrogenRepRanges.append([])  # 每个层次创建独立的列表进行储存
        for j in range(9):
            HydrogenRepRanges[i].append(ws[chr(ord("E") + j) + str(row)].value)  # 从2020年开始填入列表值
    row_accumulated += 3

    RangeNames = ["GDPRanges", "IndPorpRanges", "PopGrowRanges", "ElecSaveRanges",
                  "EVPropRanges", "PVGrowRanges", "ImpElecGrowRanges", "HydrogenRepRanges"]  # 指标名称列表
    Ranges = [GDPRanges, IndPorpRanges, PopGrowRanges, ElecSaveRanges,
              EVPropRanges, PVGrowRanges, ImpElecGrowRanges, HydrogenRepRanges]  # 指标对应范围列表
    AllRanges = dict(zip(RangeNames, Ranges))  # 创建指标名称和指标对应范围的字典（AllRanges)
    # print(AllRanges)

    return AllRanges



